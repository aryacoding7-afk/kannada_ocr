import json
from openpyxl import Workbook
from openpyxl.styles import Alignment

# -------------------------
# LOAD OCR RESULTS
# -------------------------

with open(
    "output/ocr_results.json",
    "r",
    encoding="utf-8"
) as f:

    data = json.load(f)

# -------------------------
# FIND TABLE SIZE
# -------------------------

max_row = 0
max_col = 0

for cell in data:

    max_row = max(
        max_row,
        cell["row"]
    )

    max_col = max(
        max_col,
        cell["column"]
    )

# -------------------------
# CREATE EXCEL
# -------------------------

wb = Workbook()
ws = wb.active

ws.title = "OCR Output"

# -------------------------
# WRITE CELLS
# -------------------------

for cell in data:

    row = cell["row"] + 1
    col = cell["column"] + 1

    text = cell["text"]

    ws.cell(
        row=row,
        column=col,
        value=text
    )

# -------------------------
# FORMAT
# -------------------------

for row in ws.iter_rows():

    for cell in row:

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

# -------------------------
# AUTO WIDTH
# -------------------------

for column in ws.columns:

    max_length = 0

    letter = column[0].column_letter

    for cell in column:

        try:
            value = str(cell.value)

            if len(value) > max_length:
                max_length = len(value)

        except:
            pass

    ws.column_dimensions[
        letter
    ].width = min(
        max(max_length + 2, 10),
        40
    )

# -------------------------
# SAVE
# -------------------------

output_file = "output/final.xlsx"

wb.save(output_file)

print()
print("Saved:")
print(output_file)